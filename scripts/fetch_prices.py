"""
Blue Margin — MCX Bhavcopy Scraper  [FINAL]
fetch_prices.py

All parameters confirmed from Chrome DevTools (21-Apr-2026):

  POST https://www.mcxindia.com/backpage.aspx/GetDateWiseBhavCopy
  Body : {"Date": "20260420", "InstrumentName": "ALL"}

  Response shape:
    {"d": {"Summary": {"AsOn": "...", "Count": 27778, "Status": null},
           "Data": [{"__type": "MCX.BL.Bhavcopy",
                     "Symbol": "NATURALGAS",
                     "InstrumentName": "FUTCOM",
                     "ExpiryDate": "30APR2026",
                     "OptionType": "-",
                     "StrikePrice": 0,
                     "Open": 364.35, "High": 366.35,
                     "Low": 362.15, "Close": 364.8,
                     "PreviousClose": 363.4,
                     "Volume": 2832,
                     "VolumeInThousands": "2832.000 KGS",
                     "Value": 10832.03,
                     "OpenInterest": 1895,
                     "Date": "04/20/2026",
                     "DateDisplay": "20 Apr 2026"}, ...]}}

  Selectors:
    Date display  : #txtDate          (DD/MM/YYYY)
    Date hidden   : #txtDate_hid_val  (YYYYMMDD)
    Show button   : #btnShow

Usage:
    python fetch_prices.py                            # today with look-back
    python fetch_prices.py 2026-04-20                 # one date
    python fetch_prices.py backfill 2023-01-01        # all days from date → today
    python fetch_prices.py backfill 2023-01-01 2024-12-31  # date range
    python fetch_prices.py diagnose 2026-04-20        # raw JSON dump, no DB write
"""

import asyncio
import json
import re
import sqlite3
from pathlib import Path
import sys
from datetime import datetime, timedelta
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

# Ensure UTF-8 output on Windows (cp1252 console chokes on ✓, →, … etc.)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


# ═════════════════════════════════════════════════════════════════════════════
# CONFIG
# ═════════════════════════════════════════════════════════════════════════════

HOME_URL          = "https://www.mcxindia.com/"
BHAVCOPY_URL      = "https://www.mcxindia.com/market-data/bhavcopy"
INTERCEPT_KEYWORD = "GetDateWiseBhavCopy"

DATE_INPUT_SELECTOR  = "#txtDate"
DATE_HIDDEN_SELECTOR = "#txtDate_hid_val"
SUBMIT_BTN_SELECTOR  = "#btnShow"

DATE_DISPLAY_FMT = "%d/%m/%Y"   # → visible input
DATE_HIDDEN_FMT  = "%Y%m%d"     # → hidden field + POST Date key

# Symbols we care about — new expiries appear automatically
TARGET_SYMBOLS = {"NATURALGAS", "NATGASMINI", "NATGAS"}

# Look-back for daily mode (how many trading days back to try if today has no data)
MAX_LOOKBACK = 3

# Backfill: delay between dates (seconds) — be polite to MCX servers
BACKFILL_DELAY = 2.0

# Default historical start if no date given to backfill command
DEFAULT_BACKFILL_START = "2015-01-01"

# Canonical DB path (alongside margins.db)
DEFAULT_DB_PATH = "data/prices.db"


# ═════════════════════════════════════════════════════════════════════════════
# STEALTH CONFIG  (identical to margin scraper)
# ═════════════════════════════════════════════════════════════════════════════

BROWSER_ARGS = [
    "--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage",
    "--disable-blink-features=AutomationControlled", "--disable-gpu",
    "--window-size=1920,1080",
]
CONTEXT_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "sec-ch-ua": '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "Upgrade-Insecure-Requests": "1",
}
STEALTH_SCRIPT = """
    delete Object.getPrototypeOf(navigator).webdriver;
    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
    window.chrome = {runtime: {}, loadTimes: function(){}, csi: function(){}, app: {}};
    Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
    Object.defineProperty(navigator, 'hardwareConcurrency', {get: () => 8});
"""


# ═════════════════════════════════════════════════════════════════════════════
# DATE HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def is_weekday(dt: datetime) -> bool:
    return dt.weekday() < 5   # Mon–Fri

def prev_weekday(dt: datetime) -> datetime:
    dt -= timedelta(days=1)
    while not is_weekday(dt):
        dt -= timedelta(days=1)
    return dt

def trading_days_between(start: datetime, end: datetime) -> list[datetime]:
    """Return every weekday from start to end inclusive, oldest first."""
    days, cur = [], start
    while cur <= end:
        if is_weekday(cur):
            days.append(cur)
        cur += timedelta(days=1)
    return days


# ═════════════════════════════════════════════════════════════════════════════
# SQLITE — prices stored in data/prices.db (alongside data/margins.db)
# ═════════════════════════════════════════════════════════════════════════════

COLUMNS = [
    "date", "symbol", "expiry", "instrument", "option_type", "strike_price",
    "open", "high", "low", "close", "prev_close",
    "volume_lots", "volume_kgs", "value_lacs", "open_interest",
]

CREATE_SQL = """
CREATE TABLE IF NOT EXISTS prices (
    date          TEXT    NOT NULL,
    symbol        TEXT    NOT NULL,
    expiry        TEXT    NOT NULL,
    instrument    TEXT,
    option_type   TEXT,
    strike_price  REAL,
    open          REAL,
    high          REAL,
    low           REAL,
    close         REAL,
    prev_close    REAL,
    volume_lots   INTEGER,
    volume_kgs    REAL,
    value_lacs    REAL,
    open_interest INTEGER,
    PRIMARY KEY (date, symbol, expiry)
);
CREATE INDEX IF NOT EXISTS ix_prices_date   ON prices(date);
CREATE INDEX IF NOT EXISTS ix_prices_sym    ON prices(symbol, expiry);
"""

def db_connect(db_path: str = DEFAULT_DB_PATH) -> sqlite3.Connection:
    """Open (or create) the prices SQLite DB and return a connection."""
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.executescript(CREATE_SQL)
    conn.commit()
    return conn

def dates_already_in_db(conn: sqlite3.Connection) -> set[str]:
    return {r[0] for r in conn.execute("SELECT DISTINCT date FROM prices").fetchall()}

def save_records(conn: sqlite3.Connection, records: list[dict]) -> int:
    """Upsert all records; return count of rows inserted or updated."""
    if not records:
        return 0
    placeholders = ", ".join(["?"] * len(COLUMNS))
    update_cols = [col for col in COLUMNS if col not in ("date", "symbol", "expiry")]
    update_sql = ", ".join(f"{col}=excluded.{col}" for col in update_cols)
    sql = (
        f"INSERT INTO prices ({', '.join(COLUMNS)}) VALUES ({placeholders}) "
        f"ON CONFLICT(date, symbol, expiry) DO UPDATE SET {update_sql}"
    )
    written = 0
    for r in records:
        values = tuple(r.get(c) for c in COLUMNS)
        conn.execute(sql, values)
        written += conn.execute("SELECT changes()").fetchone()[0]
    conn.commit()
    return written


# ═════════════════════════════════════════════════════════════════════════════
# PARSER + NORMALIZER
# ═════════════════════════════════════════════════════════════════════════════

def parse_response(text: str) -> list[dict]:
    """
    Parse the intercepted API body.
    Expected: {"d": {"Summary": {...}, "Data": [...]}}
    """
    text = text.strip()
    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        print(f"[parser] JSON error: {e}  first_200={text[:200]}")
        return []

    if isinstance(data, dict) and "d" in data:
        inner = data["d"]
        if isinstance(inner, str):
            inner = json.loads(inner)
        if isinstance(inner, dict):
            records = inner.get("Data") or []
            count   = inner.get("Summary", {}).get("Count", len(records))
            print(f"[parser] Count={count}  Data rows={len(records)}")
            return records if isinstance(records, list) else []
        if isinstance(inner, list):
            return inner

    if isinstance(data, list):
        return data

    print(f"[parser] Unknown shape: {list(data.keys()) if isinstance(data, dict) else type(data)}")
    return []


def _parse_vol_kgs(val) -> float | None:
    """Parse '2832.000 KGS' → 2832.0"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    m = re.match(r"[\d,]+\.?\d*", str(val).strip())
    return float(m.group().replace(",", "")) if m else None

def _to_float(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if not s or s in ("-", "N/A", "n/a", "0"):
        return None if s in ("-", "N/A", "n/a") else float(s or 0)
    try:
        return float(s)
    except ValueError:
        return None

def _to_int(val) -> int | None:
    f = _to_float(val)
    return int(f) if f is not None else None


def normalize(raw: dict, date_str: str) -> dict | None:
    """
    Map one raw API row → canonical Blue Margin record.
    All field names confirmed from DevTools Preview (Image 2).
    Returns None if symbol is not a target or row is invalid.
    """
    if not isinstance(raw, dict):
        return None

    symbol = str(raw.get("Symbol", "")).strip().upper()
    # Normalize symbol name variations
    if symbol == "NATGAS":
        symbol = "NATURALGAS"
    if symbol not in TARGET_SYMBOLS:
        return None

    expiry = str(raw.get("ExpiryDate", "")).strip()
    if not expiry or expiry == "-":
        return None

    return {
        # ── Primary key ────────────────────────────────────────────────────
        "date":          date_str,
        "symbol":        symbol,
        "expiry":        expiry,

        # ── Contract metadata ──────────────────────────────────────────────
        "instrument":    str(raw.get("InstrumentName", "")).strip(),
        "option_type":   str(raw.get("OptionType", "-")).strip(),
        "strike_price":  _to_float(raw.get("StrikePrice")),

        # ── OHLC ───────────────────────────────────────────────────────────
        "open":          _to_float(raw.get("Open")),
        "high":          _to_float(raw.get("High")),
        "low":           _to_float(raw.get("Low")),
        "close":         _to_float(raw.get("Close")),
        "prev_close":    _to_float(raw.get("PreviousClose")),

        # ── Volume & OI ────────────────────────────────────────────────────
        "volume_lots":   _to_int(raw.get("Volume")),
        "volume_kgs":    _parse_vol_kgs(raw.get("VolumeInThousands")),
        "value_lacs":    _to_float(raw.get("Value")),
        "open_interest": _to_int(raw.get("OpenInterest")),
    }


# ═════════════════════════════════════════════════════════════════════════════
# BROWSER SESSION  — single launch, multiple dates
# ═════════════════════════════════════════════════════════════════════════════

class BhavSession:
    """
    Holds a live Playwright browser + page.
    Homepage and Bhavcopy page are loaded once; subsequent fetches
    just change the date and click Show — much faster for backfill.
    """

    def __init__(self, playwright):
        self._playwright = playwright
        self.browser = None
        self.context = None
        self.page    = None
        self._ready  = False

    async def start(self):
        self.browser = await self._playwright.chromium.launch(
            headless=True, args=BROWSER_ARGS
        )
        self.context = await self.browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1920, "height": 1080},
            locale="en-US",
            timezone_id="Asia/Kolkata",
            extra_http_headers=CONTEXT_HEADERS,
        )
        await self.context.add_init_script(STEALTH_SCRIPT)
        self.page = await self.context.new_page()

        # Step 1 — Homepage for Akamai session cookie
        print("[session] Visiting homepage…")
        await self.page.goto(HOME_URL, wait_until="domcontentloaded", timeout=30_000)
        title = await self.page.title()
        if "Access Denied" in title or "blocked" in title.lower():
            raise RuntimeError("Blocked at homepage")
        await asyncio.sleep(1)

        # Step 2 — Bhavcopy page
        print("[session] Loading Bhavcopy page…")
        await self.page.goto(BHAVCOPY_URL, wait_until="domcontentloaded", timeout=30_000)
        title = await self.page.title()
        if "Access Denied" in title:
            raise RuntimeError("Blocked at Bhavcopy page")
        await self.page.wait_for_selector(DATE_INPUT_SELECTOR, timeout=15_000)
        await asyncio.sleep(2)

        self._ready = True
        print("[session] Ready ✓")

    async def fetch_date(self, date_str: str) -> list[dict]:
        """Fetch all rows for one date via direct JS fetch — no button click needed."""
        if not self._ready:
            raise RuntimeError("Session not started — call start() first")

        dt          = datetime.strptime(date_str, "%Y-%m-%d")
        date_hidden = dt.strftime(DATE_HIDDEN_FMT)

        result = await self.page.evaluate(f"""
            async () => {{
                const resp = await fetch('/backpage.aspx/GetDateWiseBhavCopy', {{
                    method: 'POST',
                    headers: {{
                        'Content-Type': 'application/json; charset=utf-8',
                        'Accept': 'application/json, text/javascript, */*; q=0.01',
                        'X-Requested-With': 'XMLHttpRequest'
                    }},
                    body: JSON.stringify({{ Date: '{date_hidden}', InstrumentName: 'ALL' }})
                }});
                return await resp.text();
            }}
        """)

        if not result:
            return []
        return parse_response(result)

    async def close(self):
        if self.browser:
            await self.browser.close()
        self._ready = False


# ═════════════════════════════════════════════════════════════════════════════
# RUNNERS
# ═════════════════════════════════════════════════════════════════════════════

async def run_daily(start_date: str | None = None, db_path: str = DEFAULT_DB_PATH):
    """
    Fetch one date (default today) with look-back buffer.
    Stops at first date that returns NATURALGAS/NATGASMINI data.
    """
    conn = db_connect(db_path)
    current = datetime.strptime(start_date, "%Y-%m-%d") if start_date else datetime.today()

    async with async_playwright() as p:
        session = BhavSession(p)
        try:
            await session.start()
        except RuntimeError as e:
            print(f"[daily] Session failed: {e}")
        
            return []

        for attempt in range(MAX_LOOKBACK + 1):
            date_str = current.strftime("%Y-%m-%d")
            print(f"\n[daily] Attempt {attempt+1}/{MAX_LOOKBACK+1}  date={date_str}")

            raw_rows   = await session.fetch_date(date_str)
            normalized = [r for raw in raw_rows if (r := normalize(raw, date_str))]

            if normalized:
                saved = save_records(conn, normalized)
                print(f"[daily] ✓ {saved} rows saved for {date_str}")
                await session.close()
            
                return normalized

            print(f"[daily] No target data for {date_str} — stepping back")
            current = prev_weekday(current)

        await session.close()
    
        print("[daily] All look-back attempts exhausted")
        return []


async def run_backfill(
    start_date: str,
    end_date: str | None = None,
    db_path: str = DEFAULT_DB_PATH,
):
    """
    Scrape every trading day from start_date → end_date (default today).
    Skips dates already in DB. Single browser session for efficiency.
    Prints a running summary as it goes.
    """
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end   = datetime.strptime(end_date, "%Y-%m-%d") if end_date else datetime.today()

    conn = db_connect(db_path)
    already_done = dates_already_in_db(conn)
    all_days     = trading_days_between(start, end)
    pending      = [d for d in all_days if d.strftime("%Y-%m-%d") not in already_done]

    total = len(all_days)
    skip  = total - len(pending)
    print(f"\n[backfill] {total} trading days in range")
    print(f"[backfill] {skip} already in DB → skipping")
    print(f"[backfill] {len(pending)} to fetch\n")

    if not pending:
        print("[backfill] Nothing to do.")
    
        return

    fetched_days  = 0
    fetched_rows  = 0
    holiday_count = 0

    async with async_playwright() as p:
        session = BhavSession(p)
        try:
            await session.start()
        except RuntimeError as e:
            print(f"[backfill] Session failed: {e}")
        
            return

        for i, dt in enumerate(pending, 1):
            date_str = dt.strftime("%Y-%m-%d")
            print(f"[backfill] {i}/{len(pending)}  {date_str}", end="  ")

            raw_rows   = await session.fetch_date(date_str)
            normalized = [r for raw in raw_rows if (r := normalize(raw, date_str))]

            if normalized:
                saved = save_records(conn, normalized)
                fetched_rows += saved
                fetched_days += 1
                syms = {r["symbol"] for r in normalized}
                expiries = {r["expiry"] for r in normalized}
                print(f"✓ {saved} rows  symbols={syms}  expiries={sorted(expiries)}")
            else:
                holiday_count += 1
                print("— no data (holiday / no trading)")

            # Polite delay between requests
            if i < len(pending):
                await asyncio.sleep(BACKFILL_DELAY)

        await session.close()


    print(f"\n[backfill] Done.")
    print(f"  Trading days with data : {fetched_days}")
    print(f"  Holidays / no data     : {holiday_count}")
    print(f"  Total rows saved       : {fetched_rows}")
    total = db_connect(db_path).execute("SELECT COUNT(*) FROM prices").fetchone()[0]
    print(f"  DB total rows          : {total}")
    print(f"  DB path                : {db_path}")


async def run_diagnose(date_str: str | None = None):
    """Print raw JSON for one date — no DB writes. Use to verify field names."""
    ds = date_str or datetime.today().strftime("%Y-%m-%d")
    async with async_playwright() as p:
        session = BhavSession(p)
        await session.start()
        raw_rows = await session.fetch_date(ds)
        await session.close()

    if not raw_rows:
        print("[diagnose] No rows returned")
        return

    # Find first NATURALGAS row if present, else first row
    sample = next(
        (r for r in raw_rows if str(r.get("Symbol", "")).strip().upper() in TARGET_SYMBOLS),
        raw_rows[0]
    )

    all_syms = sorted({str(r.get("Symbol","")).strip() for r in raw_rows})
    print(f"\n[diagnose] Total raw rows: {len(raw_rows)}")
    print(f"[diagnose] All unique symbols: {all_syms[:30]}")
    print(f"[diagnose] Target-symbol rows: {sum(1 for r in raw_rows if str(r.get('Symbol','')).strip().upper() in TARGET_SYMBOLS)}")
    print("\n[diagnose] Sample row (first target symbol or first row):")
    print(json.dumps(sample, indent=2, default=str))

    print("\n[diagnose] All keys in sample row:")
    for k, v in sample.items():
        print(f"  {k!r:30s}: {v!r}")


# ═════════════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════════════

HELP = """
Blue Margin — MCX Bhavcopy Scraper

Commands:
  (no args)                              Today with look-back, save to data/prices.db
  <YYYY-MM-DD>                           One specific date with look-back
  <YYYY-MM-DD> <db.db>                   One date, custom DB file
  backfill <start>                       All trading days from start → today
  backfill <start> <end>                 All trading days in range
  backfill <start> <end> <db.db>         Range with custom DB
  diagnose                               Dump raw JSON for today (no DB write)
  diagnose <YYYY-MM-DD>                  Dump raw JSON for specific date
  migrate                                Import blue_margin.xlsx into data/prices.db

Examples:
  python fetch_prices.py
  python fetch_prices.py 2026-04-20
  python fetch_prices.py backfill 2015-01-01
  python fetch_prices.py backfill 2023-01-01 2023-12-31
  python fetch_prices.py diagnose 2026-04-20
  python fetch_prices.py migrate
"""

def migrate_excel_to_sqlite(
    xlsx_path: str = "blue_margin.xlsx",
    db_path: str = DEFAULT_DB_PATH,
):
    """
    One-time migration: import all rows from the legacy blue_margin.xlsx
    into data/prices.db. Safe to run repeatedly — existing keys are refreshed.
    """
    import openpyxl
    if not Path(xlsx_path).exists():
        print(f"[migrate] {xlsx_path} not found — nothing to migrate.")
        return
    print(f"[migrate] Loading {xlsx_path} …")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"[migrate] Headers: {headers}")
    conn = db_connect(db_path)
    batch, written_total = [], 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        rec = dict(zip(headers, row))
        if not rec.get("date"):
            continue
        batch.append(rec)
        if len(batch) >= 500:
            written_total += save_records(conn, batch)
            batch = []
            if i % 5000 == 0:
                print(f"[migrate] … {i} rows processed, {written_total} rows written")
    if batch:
        written_total += save_records(conn, batch)
    total = conn.execute("SELECT COUNT(*) FROM prices").fetchone()[0]
    print(f"[migrate] Done.  Wrote {written_total} rows → DB total: {total}")
    print(f"[migrate] DB: {db_path}")
    conn.close()
    wb.close()


if __name__ == "__main__":
    args = sys.argv[1:]

    if not args or args[0] in ("-h", "--help", "help"):
        print(HELP)
        sys.exit(0)

    # ── migrate ──────────────────────────────────────────────────────────────
    if args[0] == "migrate":
        xlsx = args[1] if len(args) > 1 else "blue_margin.xlsx"
        db   = args[2] if len(args) > 2 else DEFAULT_DB_PATH
        migrate_excel_to_sqlite(xlsx, db)
        sys.exit(0)

    # ── diagnose ─────────────────────────────────────────────────────────────
    if args[0] == "diagnose":
        asyncio.run(run_diagnose(args[1] if len(args) > 1 else None))
        sys.exit(0)

    # ── backfill ─────────────────────────────────────────────────────────────
    if args[0] == "backfill":
        if len(args) < 2:
            print("Usage: python fetch_prices.py backfill <start_date> [end_date] [db_path]")
            sys.exit(1)
        start_arg = args[1]
        end_arg   = args[2] if len(args) > 2 and not args[2].endswith(".db") else None
        db_arg    = next((a for a in args[2:] if a.endswith(".db")), DEFAULT_DB_PATH)
        asyncio.run(run_backfill(start_arg, end_arg, db_arg))
        sys.exit(0)

    # ── daily (default) ───────────────────────────────────────────────────────
    date_arg = args[0] if args else None
    db_arg   = next((a for a in args if a.endswith(".db")), DEFAULT_DB_PATH)
    results  = asyncio.run(run_daily(start_date=date_arg, db_path=db_arg))

    if results:
        print(f"\n✓ {len(results)} records  date={results[0]['date']}")
        for r in results[:6]:
            print(f"  {r['symbol']:12s} {r['expiry']:12s}  "
                  f"O={r['open']}  H={r['high']}  L={r['low']}  C={r['close']}  "
                  f"Lots={r['volume_lots']}  OI={r['open_interest']}")
        if len(results) > 6:
            print(f"  … and {len(results)-6} more.")
    else:
        print("✗ No data retrieved.")
        sys.exit(1)
